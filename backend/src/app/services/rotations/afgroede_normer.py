"""Afgrødenormer (udbytte, N-norm, M/W/MP/WP) fra Bilag 1-mastertabellen.

Porteret verbatim fra c:\\plantperform-nles\\src\\afgroede_normer.py — samme
logik, sti tilpasset DST2's database/data/raw/ANGJ-data/.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[4]  # .../backend
EXCEL_PATH = (
    _ROOT / "database" / "data" / "raw" / "ANGJ-data"
    / "PlantPerform_master_afgroedenormer_opdateret_fra_hoeringsmateriale_Bilag_1_1_2027.xlsx"
)

# Kartofler anvender M2 (Vårsæd) i NLES5 — NUAR AU-anbefaling gældende fra 2027-regulering.
# W ændres IKKE (typisk W3). Reglen gælder altid og overskriver eventuelle andre M-værdier.
KARTOFFEL_KODER: frozenset = frozenset({149, 150, 151, 152, 154, 155, 156})


def apply_kartoffel_regel(sample: dict) -> dict:
    """Returnerer sample med M=2 for kartoffelkoder; alle andre felter uændret.

    Kartoffelkoder: 149, 150, 151, 152, 154, 155, 156.
    Reglen overskriver M uanset hvad brugeren har valgt i UI'en.
    W ændres ikke.
    """
    if sample.get("crop_code") in KARTOFFEL_KODER:
        return {**sample, "M": 2}
    return sample


_VANDING_PRIORITY = {
    True: ["Vandet", "Ikke særskilt vanding", "Uvandet"],
    False: ["Uvandet", "Ikke særskilt vanding", "Vandet"],
}


def _parse_jb_set(match_type, jb_values_str):
    """Expand JB_match_type + JB_værdier into a set of integer JB numbers."""
    s = str(jb_values_str).strip() if jb_values_str is not None else ""
    if not s or s.lower() == "none":
        return set()
    mt = str(match_type).strip().lower() if match_type else ""
    try:
        if mt == "plus":
            return {int(x) for x in s.split(";")}
        elif mt == "til":
            parts = s.split("-")
            return set(range(int(parts[0]), int(parts[1]) + 1))
        elif mt == "enkelt":
            return {int(s)}
    except (ValueError, IndexError):
        pass
    return set()


def _parse_jb_set_infer(jb_values_str):
    """Parse JB values without an explicit match_type column (infer from format)."""
    s = str(jb_values_str).strip() if jb_values_str is not None else ""
    if not s or s.lower() == "none":
        return set()
    try:
        if ";" in s:
            return {int(x) for x in s.split(";")}
        elif "-" in s:
            parts = s.split("-")
            return set(range(int(parts[0]), int(parts[1]) + 1))
        else:
            return {int(s)}
    except (ValueError, IndexError):
        return set()


def _to_float(val):
    """Parse numeric value; handles 'X/Y' strings by taking first part."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        s = str(val).strip()
        if "/" in s:
            try:
                return float(s.split("/")[0])
            except ValueError:
                pass
        return None


@lru_cache(maxsize=1)
def _load_lang_lookup():
    """Load Lang_lookup sheet → dict: (crop_code, jb_nr, vanding) → row_data.

    Column indices (0-based):
      0  Afgrødekode
      1  Afgrøde
      4  JB_gruppe
      5  JB_match_type   (plus / til / enkelt)
      6  JB_værdier      (e.g. "1;3", "1-4", "11")
      7  Vanding         (Uvandet / Vandet / Ikke særskilt vanding)
      9  Udbytteenhed
      10 Udbyttenorm
      11 Udbyttenorm_alt_ikke_korn
      13 N_norm_kgN_ha
      14 P_norm_kgP_ha
      15 Forfrugtsværdi_kgN_ha
      16 Indregn_forfrugtsværdi_i_N_norm
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Lang_lookup"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    lookup = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            crop_code = int(row[0])
        except (ValueError, TypeError):
            continue

        jb_set = _parse_jb_set(row[5], row[6])
        if not jb_set:
            continue

        vanding = str(row[7]).strip() if row[7] else ""
        fv_raw = _to_float(row[15])

        row_data = {
            "afgroede": str(row[1]).strip() if row[1] else "",
            "jb_gruppe": str(row[4]).strip() if row[4] else "",
            "vanding": vanding,
            "udbytteenhed": str(row[9]).strip() if row[9] else "",
            "udbyttenorm": _to_float(row[10]),
            "udbyttenorm_alt": _to_float(row[11]),
            "n_norm": _to_float(row[13]),
            "p_norm": _to_float(row[14]),
            "forfrugtsvaerdi": fv_raw if fv_raw is not None else 0.0,
            "indregn_ffv": str(row[16]).strip().lower() == "ja" if row[16] else False,
        }

        for jb_nr in jb_set:
            lookup[(crop_code, jb_nr, vanding)] = row_data

    return lookup


@lru_cache(maxsize=1)
def _load_nfix_lookup():
    """Load N_fixering_lookup sheet → dict: (crop_code, jb_nr, vanding) → nfix_kgN_ha.

    N_fixering_lookup columns (0-based):
      0  Nfix_lookup_key  (e.g. "7|JB 1 + 3")
      1  Afgrødekode
      2  Afgrøde
      3  Nfix_klasse
      4  JB_gruppe
      5  JB_værdier       (e.g. "1;3", "1-4", "11")
      6  Vanding
      7  Jordtypegruppe
      8  Nfix_kgN_ha
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["N_fixering_lookup"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    lookup = {}
    for row in rows[1:]:
        if not row[1]:
            continue
        try:
            crop_code = int(row[1])
        except (ValueError, TypeError):
            continue

        jb_set = _parse_jb_set_infer(row[5])
        if not jb_set:
            continue

        vanding = str(row[6]).strip() if row[6] else ""
        nfix = _to_float(row[8])
        if nfix is None:
            continue

        for jb_nr in jb_set:
            lookup[(crop_code, jb_nr, vanding)] = nfix

    return lookup


def lookup_norm(crop_code, jb_nr, irrigated=False):
    """Return norm data dict for (crop_code, jb_nr, irrigated), or None if not found.

    Priority: Vandet/Uvandet exact match → Ikke særskilt vanding → other.
    """
    if crop_code is None or jb_nr is None:
        return None
    lut = _load_lang_lookup()
    for vanding in _VANDING_PRIORITY[bool(irrigated)]:
        key = (crop_code, jb_nr, vanding)
        if key in lut:
            return lut[key]
    return None


def lookup_nfix(crop_code, jb_nr, irrigated=False):
    """Return Nfix_kgN_ha for (crop_code, jb_nr, irrigated), or 0.0 if not found.

    Nfix er biologisk kvælstoffiksering til NUAR F0/F1/F2 — NOT forfrugtsværdi.
    """
    if crop_code is None or jb_nr is None:
        return 0.0
    lut = _load_nfix_lookup()
    for vanding in _VANDING_PRIORITY[bool(irrigated)]:
        key = (crop_code, jb_nr, vanding)
        if key in lut:
            return lut[key]
    return 0.0


@lru_cache(maxsize=1)
def _load_nuar_koder():
    """Load NUAR_koder sheet → dict: crop_code → {naam, M, W, WC, MP, WP, *_ambig}.

    Column indices (0-based):
      0  AfgroedeKode
      1  Naam
      2  M       (Hovedafgrøde)
      3  W       (Vinterplantedække)
      4  WC      (Efterårsoptag)
      5  MP      (Forfrugtskategori)
      6  WP      (Forfrugts vinterdække)
      7  M_ambig  (0/1)
      8  W_ambig
      9  WC_ambig
      10 MP_ambig
      11 WP_ambig
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["NUAR_koder"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    lookup = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            code = int(row[0])
        except (ValueError, TypeError):
            continue

        def _iv(val):
            if val is None:
                return None
            try:
                return int(val)
            except (ValueError, TypeError):
                return None

        lookup[code] = {
            "navn": str(row[1]).strip() if row[1] else "",
            "M": _iv(row[2]),
            "W": _iv(row[3]),
            "WC": _iv(row[4]),
            "MP": _iv(row[5]),
            "WP": _iv(row[6]),
            "M_ambig": bool(row[7]),
            "W_ambig": bool(row[8]),
            "WC_ambig": bool(row[9]),
            "MP_ambig": bool(row[10]),
            "WP_ambig": bool(row[11]),
        }
    return lookup


def lookup_crop_params(crop_code):
    """Return NUAR-parametre {naam, M, W, WC, MP, WP, *_ambig} for crop_code, or {}."""
    if crop_code is None:
        return {}
    return _load_nuar_koder().get(crop_code, {})


def crop_names_from_normer():
    """Return {crop_code: naam} from NUAR_koder (falls back to Lang_lookup)."""
    koder = _load_nuar_koder()
    if koder:
        return {code: info["navn"] for code, info in koder.items()}
    lut = _load_lang_lookup()
    names = {}
    for (code, _jb, _v), data in lut.items():
        if code not in names:
            names[code] = data["afgroede"]
    return names
